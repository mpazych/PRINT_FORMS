from openpyxl import load_workbook 
from openpyxl import Workbook
from read_xl import reading_colums, reading_row

from os import listdir

from layout import pdf_creator, sheet_title, table_header_epic, table_line_epic, \
                   table_header_line, sheet_footer
from layout import up_table_line, table_footer
from fpdf import FPDF

from in_words import num_in_words

"""Reading data from delivery note, columns numbers are defined by it's name"""

del_note = '2515998375'
wb = load_workbook(str('./Excel/' + del_note + '.xlsx'))
ws = wb.active


HEADER_ROW = reading_row (ws, 1)

customer_num  = ws.cell(row = 2,column = HEADER_ROW.index('ПолучатМтр') + 1).value
customer_name = ws.cell(row = 2,column = HEADER_ROW.index('Имя 1')      + 1).value
order_n       = ws.cell(row = 2,column = HEADER_ROW.index('Cust.PO No') + 1).value
doc_date      = str(ws.cell(row = 2,column = HEADER_ROW.index('Д/докум.')   + 1).value)
crit          = HEADER_ROW.index('АЕИ') + 1 # column num with BOX QTY A02 criteria
pc = 'PC' # could be also 'ST'
box = 'BOX' # could be also 'KAR'

EAN      = reading_colums (ws, HEADER_ROW.index('Код EAN/UPC') + 1       , crit, pc)
REF      = reading_colums (ws, HEADER_ROW.index('Название стандарта') + 1, crit, pc)
REF_CUST = reading_colums (ws, HEADER_ROW.index('Cust.Mat.No') + 1       , crit, pc)
DESC     = reading_colums (ws, HEADER_ROW.index('Название') + 1          , crit, pc)
CMMF     = reading_colums (ws, HEADER_ROW.index('Материал') + 1          , crit, pc)
QTY      = reading_colums (ws, HEADER_ROW.index('Deliv. Q-ty') + 1       , crit, pc)
AMOUNT   = reading_colums (ws, HEADER_ROW.index('СтоимНетт') + 1         , crit, pc)
BOX      = reading_colums (ws, HEADER_ROW.index('Numerator') + 1         , crit, box)

print (len(REF), len(REF_CUST))

sum_amount = 0.00
sum_qty = 0

for i in range (len(QTY)):
    sum_amount = sum_amount + AMOUNT[i]
    sum_qty = sum_qty + QTY[i]

""" testing block, to be changed"""
sum = round (sum_amount*1.2,2)
vat = str(format(round (sum / 6, 2),',.2f')) + ' грн.'
result = ""
sum_in_word = num_in_words (sum)


""" Reading GSU and customer data for document header """

wb = load_workbook ('./Excel/Clients.xlsx')
ws = wb.active

GSU_HEADER, DOC_HEADER = [], []
CLIENTS = reading_row (ws, 1)

for i in range (2,7):
    DOC_HEADER.append (ws.cell(row = i,column = CLIENTS.index(customer_name)+1).value)
    GSU_HEADER.append (ws.cell(row = i,column = CLIENTS.index('GSU')+1).value)


"""creating pdf file"""

"""loading FPDF """
pdf = FPDF('L','mm','A4')
pdf_creator (pdf)

"""creating sheet title and table header"""
sheet_title (GSU_HEADER, DOC_HEADER, order_n, pdf)
up_table_line (del_note, doc_date, pdf)
table_header_epic (pdf)

"""filling the table by lines"""
for i in range (len(CMMF)):
    table_line_epic (str(REF[i]), str(REF_CUST[i]), str(i + 1), 
                     str(EAN[i]), QTY[i],str(DESC[i]), AMOUNT[i], BOX[i], pdf)
    if pdf.get_y() > 180:
        pdf.cell(276, 0, '','T',1,'L') # page bottom line
        up_table_line (del_note, doc_date, pdf)
        pdf.add_page()
        table_header_line (pdf)
        
table_footer (sum_qty, sum_amount, pdf)

pdf.cell(275, 3, '',0,1,'L') # empty line


pdf.cell(30, 4, 'Всього на суму із ПДВ:',0,0,'R')
pdf.cell(250, 4, sum_in_word ,0 ,1 ,'L')
pdf.cell(30, 4, 'ПДВ:',0,0,'R')
pdf.cell(30, 4, vat ,0 ,1 ,'L')

pdf.cell(70, 2, '',0,1,'L') # empty line

sheet_footer (pdf)

y = pdf.get_y()
print ('y= ',y)


pdf.output(str('Outputs\\' + del_note + '.pdf' ), 'F')
