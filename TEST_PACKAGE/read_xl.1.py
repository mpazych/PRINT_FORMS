
#import FPDF_layout
from openpyxl import load_workbook 
from openpyxl import Workbook

from fpdf import FPDF
pdf = FPDF('L','mm','A4')
pdf.add_page()
pdf.add_font('Arial2', '', './fonts/arial.ttf', uni=True) 
pdf.add_font('Arial2', 'I', './fonts/ariali.ttf', uni=True) 
pdf.set_font('Arial2', '', 8) # fpdf.set_font(family, style = '', size = 0)

wb = load_workbook('./Excel/2515921043.xlsx') #fpdf class
#wb = Workbook('./Excel/2515921043.xlsx') # why is not woriking?
ws = wb.active #fpdf class

def table_header():  #шапка таблицы
    """Первая строка"""
    pdf.cell(12, 5, 'Номер','TLR',0,'C') # Номер пп
    pdf.cell(80, 5, 'Товар', 1, 0, 'C') # Товар
    pdf.cell(15, 5, 'Од.', 'TLR', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, 'Вид', 'TLR', 0, 'C')  # Вид упаковки
    pdf.cell(30, 5, 'Кількість', 1, 0, 'C')  # Кiлькiсть
    pdf.cell(15, 5, 'Маса', 'TLR', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, 'Маса', 'TLR', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, 'Ціна,', 'TLR', 0, 'C')  # Ціна
    pdf.cell(25, 5, 'Сума без', 'TLR', 0, 'C')  # Сумма без ПДВ
    pdf.cell(25, 5, 'ПДВ', 1, 0, 'C')  # ПДВ
    pdf.cell(25, 5, 'Сума з', 'TLR', 1, 'C')  # Сумма с ПДВ
    """Вторая строка"""
    pdf.cell(12, 5, 'п/п','LR',0,'C') # Номер пп
    pdf.cell(60, 5, 'Найменування', 'LR', 0, 'C') # Наименование
    pdf.cell(20, 5, 'Код', 'LR', 0, 'C') # Код артикула
    pdf.cell(15, 5, 'виміру', 'LR', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, 'упако-', 'LR', 0, 'C')  # Вид упаковки
    pdf.cell(15, 5, 'в', 'LR', 0, 'C')  # в одному місці
    pdf.cell(15, 5, 'місць', 'LR', 0, 'C')  # місць
    pdf.cell(15, 5, 'брутто', 'LR', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, 'нетто', 'LR', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, 'грн.коп', 'LR', 0, 'C')  # Ціна
    pdf.cell(25, 5, 'урахування', 'LR', 0, 'C')  # Сумма без ПДВ
    pdf.cell(10, 5, 'Ставка', 'LR', 0, 'C')  # ПДВ
    pdf.cell(15, 5, 'Сума,', 'LR', 0, 'C')  # ПДВ
    pdf.cell(25, 5, 'урахуванням', 'LR', 1, 'C')  # Сумма с ПДВ
    """Третья строка"""
    pdf.cell(12, 5, '','LR',0,'C') # Номер пп
    pdf.cell(60, 5, 'характеристики, сорт артикулу', 'LR', 0, 'C') # Наименование
    pdf.cell(20, 5, 'артикулу', 'LR', 0, 'C') # Код артикула
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, 'вки', 'LR', 0, 'C')  # Вид упаковки
    pdf.cell(15, 5, 'одному', 'LR', 0, 'C')  # в одному місці
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # місць
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, '', 'LR', 0, 'C')  # Ціна
    pdf.cell(25, 5, 'ПДВ,', 'LR', 0, 'C')  # Сумма без ПДВ
    pdf.cell(10, 5, '%', 'LR', 0, 'C')  # ПДВ
    pdf.cell(15, 5, 'грн.коп', 'LR', 0, 'C')  # ПДВ
    pdf.cell(25, 5, 'ПДВ,', 'LR', 1, 'C')  # Сумма с ПДВ
    """Четвертая строка"""
    pdf.cell(12, 5, '','LR',0,'C') # Номер пп
    pdf.cell(60, 5, '', 'LR', 0, 'C') # Наименование
    pdf.cell(20, 5, '', 'LR', 0, 'C') # Код артикула
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Вид упаковки
    pdf.cell(15, 5, 'місці', 'LR', 0, 'C')  # в одному місці
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # місць
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, '', 'LR', 0, 'C')  # Ціна
    pdf.cell(25, 5, 'грн.коп', 'LR', 0, 'C')  # Сумма без ПДВ
    pdf.cell(10, 5, '', 'LR', 0, 'C')  # ПДВ
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # ПДВ
    pdf.cell(25, 5, 'грн.коп', 'LR', 1, 'C')  # Сумма с ПДВ
    """Пятая строка"""
    pdf.cell(12, 5, '1','LRB',0,'C') # Номер пп
    pdf.cell(60, 5, '2', 'LRB', 0, 'C') # Наименование
    pdf.cell(20, 5, '3', 'LRB', 0, 'C') # Код артикула
    pdf.cell(15, 5, '4', 'LRB', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, '5', 'LRB', 0, 'C')  # Вид упаковки
    pdf.cell(15, 5, '6', 'LRB', 0, 'C')  # в одному місці
    pdf.cell(15, 5, '7', 'LRB', 0, 'C')  # місць
    pdf.cell(15, 5, '8', 'LRB', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, '9', 'LRB', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, '10', 'LRB', 0, 'C')  # Ціна
    pdf.cell(25, 5, '11', 'LRB', 0, 'C')  # Сумма без ПДВ
    pdf.cell(10, 5, '12', 'LRB', 0, 'C')  # ПДВ
    pdf.cell(15, 5, '13', 'LRB', 0, 'C')  # ПДВ
    pdf.cell(25, 5, '14', 'LRB', 1, 'C')  # Сумма с ПДВ

def table_line(a): # строка таблицы
    pdf.cell(12, 5, str(a),'LR',0,'C') # Номер пп
    pdf.cell(60, 5, DESC [a], 'LR', 0, 'L') # Наименование
    pdf.cell(20, 5, CMMF [a], 'LR', 0, 'C') # Код артикула
    pdf.cell(15, 5, 'шт', 'LR', 0, 'C')  # Од. вимiру
    pdf.cell(15, 5, 'короб', 'LR', 0, 'C')  # Вид упаковки
    pdf.cell(15, 5, '**', 'LR', 0, 'C')  # в одному місці
    pdf.cell(15, 5, str(QTY [a]), 'LR', 0, 'C')  # місць
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса брутто
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # Масса нетто
    pdf.cell(25, 5, '', 'LR', 0, 'C')  # Ціна
    pdf.cell(25, 5, str(NET [a]), 'LR', 0, 'C')  # Сумма без ПДВ
    pdf.cell(10, 5, '', 'LR', 0, 'C')  # ПДВ
    pdf.cell(15, 5, '', 'LR', 0, 'C')  # ПДВ
    pdf.cell(25, 5, '', 'LR', 1, 'C')  # Сумма с ПДВ

# Header_line = []
# j, i = 1, 1
# while ws.cell(row=i,column=j).value != None:
#     Header_line.append (ws.cell(row=i,column=j).value)
#     j +=1
# print (Header_line[:], len(Header_line))

CMMF = []
DESC = []
QTY = []
NET = []
c_cmmf = 5
c_desc = 6
c_qty = 7
c_net = 9

i = 1
while ws.cell(row=i,column=c_cmmf).value != None:
    CMMF.append (ws.cell(row=i,column=c_cmmf).value)
    DESC.append (ws.cell(row=i,column=c_desc).value)
    QTY.append (ws.cell(row=i,column=c_qty).value)
    NET.append (ws.cell(row=i,column=c_net).value)
    i +=1

table_header()

a = 1
while a <= len (CMMF)-2:
    table_line (a)
    a +=1

pdf.output('Outputs\FPDF_FILE.pdf', 'F')
#print (len(CMMF),len(DESC),len(QTY))

