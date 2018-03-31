
from openpyxl import load_workbook 
from openpyxl import Workbook


def reading_data(REF, REF_CUST, DESC, CMMF, EAN, QTY, AMOUNT, del_note):
    "Reading data from the delivery note"
    wb = load_workbook(str('./Excel/'+del_note+'.xlsx'))
    ws = wb.active 
   
  
    i = 2
    while ws.cell(row = i,column = 1).value != None or i == 200:
        REF.append (ws.cell(row = i, column = 20).value)
        REF_CUST.append (ws.cell(row = i, column = 64).value)
        DESC.append (ws.cell(row = i, column = 43).value)
        CMMF.append (ws.cell(row = i, column = 19).value)
        EAN.append (ws.cell(row = i, column = 21).value)
        QTY.append (ws.cell(row = i, column = 22).value)
        AMOUNT.append (ws.cell(row = i, column = 68).value)
        i +=3
    #return DESC,CMMF,EAN,QTY

def reading_doc_header (GSU_HEADER, DOC_HEADER, CLIENT_LIST, client):
    "Reading data for document header"
    wb = load_workbook('./Excel/Clients.xlsx')
    ws = wb.active 
    i = 2
    while ws.cell(row=1,column=i).value != None:
        CLIENT_LIST.append (ws.cell(row=1,column=i).value)
        i +=1
    client_index = CLIENT_LIST.index(client)

    for i in range (2,7):
        DOC_HEADER.append (ws.cell(row=i,column=client_index+2).value)
        GSU_HEADER.append (ws.cell(row=i,column=CLIENT_LIST.index('GSU')+2).value)
    # return GSU_HEADER, DOC_HEADER, CLIENT_LIST

def reading_cell (row,column,del_note):
    wb = load_workbook(str('./Excel/'+del_note+'.xlsx'))
    ws = wb.active #fpdf class
    return ws.cell(row = row,column = column).value

""" FINAL FORMULAS"""

def reading_colums (ws, column_num, crit, crit_value):
    "Reading column in the list, max nb of items 2000, till the end of the 1st column in excel"
    AAA = []
    i = 2
    while ws.cell(row = i,column = 1).value != None or i == 2000:
        if ws.cell(row = i,column = crit).value == crit_value:
            if ws.cell(row = i,column = column_num).value == None:
                AAA.append (" ")
            else:
                AAA.append (ws.cell(row = i, column = column_num).value)
        i +=1 
    return AAA

def reading_row (ws,row_num):
    "Reading row in the list, max nb of items 200, till the end of header row"
    AAA = []
    i = 1
    while ws.cell(row = 1,column = i).value != None or i == 200:
        if ws.cell(row = row_num,column = i).value == None:
            AAA.append (" ")
        else:
            AAA.append (ws.cell(row = row_num,column = i).value)
        i +=1 
    return AAA



   

